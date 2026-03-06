"""
File upload parsers — CSV and XLSX support for suppliers, demand, and P2P events.
"""
from __future__ import annotations

import csv
import io
import json
from typing import Any

from fastapi import UploadFile


def _read_csv(content: bytes) -> list[dict]:
    """Parse CSV bytes into list of dicts."""
    text = content.decode("utf-8-sig")  # handle BOM
    reader = csv.DictReader(io.StringIO(text), delimiter=None)
    # auto-detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:2048])
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    except csv.Error:
        reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _read_xlsx(content: bytes) -> list[dict]:
    """Parse XLSX bytes into list of dicts (first sheet, first row = headers)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed — cannot read XLSX files")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h).strip() for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]


async def _read_file(file: UploadFile) -> list[dict]:
    """Read uploaded file and return rows as dicts."""
    content = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return _read_xlsx(content)
    return _read_csv(content)


def _float(val: Any, default: float = 0.0) -> float:
    if val is None or str(val).strip() == "":
        return default
    return float(val)


def _str(val: Any, default: str = "") -> str:
    if val is None:
        return default
    return str(val).strip()


# ---------------------------------------------------------------------------
# Suppliers parser
# ---------------------------------------------------------------------------
# Expected columns (flexible naming):
#   supplier_id, name, unit_cost, logistics_cost, lead_time_days,
#   compliance_score, esg_score, min_order_qty, max_capacity, served_regions

_SUPPLIER_COL_MAP = {
    "supplier_id": ["supplier_id", "id", "sup_id", "kod_dostawcy"],
    "name": ["name", "nazwa", "supplier_name", "dostawca"],
    "unit_cost": ["unit_cost", "koszt", "cena", "cost", "cena_jedn"],
    "logistics_cost": ["logistics_cost", "koszt_logistyki", "transport", "log_cost"],
    "lead_time_days": ["lead_time_days", "lead_time", "czas_dostawy", "dni"],
    "compliance_score": ["compliance_score", "compliance", "sla", "zgodnosc"],
    "esg_score": ["esg_score", "esg", "ekologia", "reliability", "niezawodnosc"],
    "min_order_qty": ["min_order_qty", "moq", "min_qty", "min_zamowienie"],
    "max_capacity": ["max_capacity", "capacity", "pojemnosc", "maks_pojemnosc"],
    "served_regions": ["served_regions", "regions", "regiony", "region"],
}


def _find_col(row_keys: list[str], aliases: list[str]) -> str | None:
    """Find a column name from a row using alias matching."""
    lower_keys = {k.lower().strip(): k for k in row_keys}
    for alias in aliases:
        if alias in lower_keys:
            return lower_keys[alias]
    return None


def _map_row(row: dict, col_map: dict[str, list[str]]) -> dict[str, Any]:
    """Map a row using flexible column name aliases."""
    keys = list(row.keys())
    mapped = {}
    for target, aliases in col_map.items():
        col = _find_col(keys, aliases)
        if col is not None:
            mapped[target] = row[col]
    return mapped


async def parse_suppliers_file(file: UploadFile) -> list[dict]:
    """Parse uploaded suppliers file into list of supplier dicts."""
    raw = await _read_file(file)
    result = []
    for row in raw:
        m = _map_row(row, _SUPPLIER_COL_MAP)
        if not m.get("supplier_id") or not m.get("name"):
            continue

        regions = m.get("served_regions", "")
        if isinstance(regions, str):
            # Try JSON array, then comma-separated
            try:
                regions = json.loads(regions)
            except (json.JSONDecodeError, TypeError):
                regions = [r.strip() for r in regions.split(",") if r.strip()]

        result.append({
            "supplier_id": _str(m["supplier_id"]),
            "name": _str(m["name"]),
            "unit_cost": _float(m.get("unit_cost")),
            "logistics_cost": _float(m.get("logistics_cost")),
            "lead_time_days": _float(m.get("lead_time_days")),
            "compliance_score": _float(m.get("compliance_score")),
            "esg_score": _float(m.get("esg_score")),
            "min_order_qty": _float(m.get("min_order_qty")),
            "max_capacity": _float(m.get("max_capacity"), 1),
            "served_regions": regions if isinstance(regions, list) and regions else ["PL"],
        })
    return result


# ---------------------------------------------------------------------------
# Demand parser
# ---------------------------------------------------------------------------

_DEMAND_COL_MAP = {
    "product_id": ["product_id", "id", "indeks", "produkt", "product", "idx"],
    "demand_qty": ["demand_qty", "qty", "ilosc", "demand", "zapotrzebowanie", "quantity"],
    "destination_region": ["destination_region", "region", "region_docelowy", "dest"],
}


async def parse_demand_file(file: UploadFile) -> list[dict]:
    """Parse uploaded demand file."""
    raw = await _read_file(file)
    result = []
    for row in raw:
        m = _map_row(row, _DEMAND_COL_MAP)
        if not m.get("product_id") or not m.get("demand_qty"):
            continue
        result.append({
            "product_id": _str(m["product_id"]),
            "demand_qty": _float(m["demand_qty"]),
            "destination_region": _str(m.get("destination_region"), "PL"),
        })
    return result


# ---------------------------------------------------------------------------
# P2P Events parser
# ---------------------------------------------------------------------------

_P2P_COL_MAP = {
    "case_id": ["case_id", "case", "nr_sprawy", "req", "id_procesu"],
    "activity": ["activity", "czynnosc", "aktywnosc", "step", "krok", "event"],
    "timestamp": ["timestamp", "czas", "data", "time", "datetime"],
    "resource": ["resource", "zasob", "uzytkownik", "user", "kto"],
    "cost": ["cost", "koszt", "wartosc", "value"],
}


async def parse_p2p_events_file(file: UploadFile) -> list[dict]:
    """Parse uploaded P2P event log file."""
    raw = await _read_file(file)
    result = []
    for row in raw:
        m = _map_row(row, _P2P_COL_MAP)
        if not m.get("case_id") or not m.get("activity") or not m.get("timestamp"):
            continue
        result.append({
            "case_id": _str(m["case_id"]),
            "activity": _str(m["activity"]),
            "timestamp": _str(m["timestamp"]),
            "resource": _str(m.get("resource")),
            "cost": _float(m.get("cost")),
        })
    return result
